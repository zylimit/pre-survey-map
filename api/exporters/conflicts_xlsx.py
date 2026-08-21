"""把 session 里的 conflicts 列表导成 Excel。

Spec F5：列含 类型 / 名称 / 库里现状 / 导入文件 / 来源文件名 / 双方所有字段对照。

实现成四个 sheet：Site / Road / Lessor / Area Conflicts（按 kind 分桶，
只产非空桶；#51 review CRITICAL-1 修复——area 冲突原先无桶被静默丢弃，纯 area
冲突时输出「无冲突」谎报；同根缺陷的 F20 road 变体同批修复——road 冲突亦无桶）。
列布局：A 类型 | B 名称 | C 来源文件 | D 起每个字段一对：「字段 [DB]」「字段 [新]」。
字段集 = 该类型所有冲突行里出现过的字段（核心列 + extras 并集）。
"""

import io
import json
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


SITE_CORE = ["site_id", "option", "project", "site_status", "lati", "longi"]
LESSOR_CORE = ["fid", "lessor_name", "lessor_category", "relationship"]
# #51：area 核心列 = name（去重键）+ operator（盖戳列）
AREA_CORE = ["name", "operator"]
# F20：road 核心列 = property（去重键；空 Property 无身份、永不冲突，故无空值场景）
ROAD_CORE = ["property"]

# extras 里要排除的、已经在核心列里出现过的键（避免重复）
SITE_EXTRA_DROP = {"SITE ID", "OPTION", "PROJECT", "SITE STATUS", "LATI", "LONGI"}
LESSOR_EXTRA_DROP = {"fid", "Lessor Name", "Lessor Category", "Lessor Cagegory", "Relationship"}
# 同 parsers/kml.py 的 _AREA_CORE 口径：name/operator 大小写写法都排除，不回灌 extras
AREA_EXTRA_DROP = {"name", "Name", "operator", "OPERATOR"}
# 同 parsers/kml.py 的 _ROAD_CORE 口径：KML 里是 "Property"、DB 列是 "property"，都排除
ROAD_EXTRA_DROP = {"Property", "property"}

_KIND_CONF = {
    "site": (SITE_CORE, SITE_EXTRA_DROP),
    "lessor": (LESSOR_CORE, LESSOR_EXTRA_DROP),
    "area": (AREA_CORE, AREA_EXTRA_DROP),
    "road": (ROAD_CORE, ROAD_EXTRA_DROP),
}

HEADER_FILL = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
DB_FILL = PatternFill(start_color="FFEAF4FF", end_color="FFEAF4FF", fill_type="solid")
NEW_FILL = PatternFill(start_color="FFFFF8E1", end_color="FFFFF8E1", fill_type="solid")


def _parse_extras(extras: Any) -> dict[str, Any]:
    if extras is None:
        return {}
    if isinstance(extras, str):
        try:
            return json.loads(extras)
        except json.JSONDecodeError:
            return {}
    if isinstance(extras, dict):
        return extras
    return {}


def _collect_fields(conflicts: list[dict[str, Any]], core: list[str], drop: set[str]) -> list[str]:
    """字段顺序：核心列 + sorted(extras 并集 - drop)。"""
    extras = set()
    for c in conflicts:
        for row in (c.get("existing", {}), c.get("incoming", {})):
            for k in _parse_extras(row.get("extras")).keys():
                if k in drop:
                    continue
                extras.add(k)
    return core + sorted(extras)


def _value(row: dict[str, Any], field: str, core: list[str]) -> Any:
    if field in core:
        return row.get(field)
    return _parse_extras(row.get("extras")).get(field, "")


def _write_sheet(ws, kind: str, conflicts: list[dict[str, Any]]):
    core, drop = _KIND_CONF[kind]
    fields = _collect_fields(conflicts, core, drop)

    # ---- 表头 ----
    ws.cell(1, 1, "类型").font = Font(bold=True)
    ws.cell(1, 2, "名称").font = Font(bold=True)
    ws.cell(1, 3, "来源文件名").font = Font(bold=True)
    col = 4
    for f in fields:
        cell_db = ws.cell(1, col, f"{f} [DB]")
        cell_db.font = Font(bold=True)
        cell_db.fill = DB_FILL
        cell_new = ws.cell(1, col + 1, f"{f} [新]")
        cell_new.font = Font(bold=True)
        cell_new.fill = NEW_FILL
        col += 2

    for c in range(1, ws.max_column + 1):
        ws.cell(1, c).alignment = Alignment(horizontal="center", vertical="center")
    for c in range(1, 4):
        ws.cell(1, c).fill = HEADER_FILL

    # ---- 数据行 ----
    for row_idx, c in enumerate(conflicts, start=2):
        ws.cell(row_idx, 1, c["kind"])
        ws.cell(row_idx, 2, c["name"])
        ws.cell(row_idx, 3, c.get("source_file") or "")
        col = 4
        existing = c.get("existing", {})
        incoming = c.get("incoming", {})
        for f in fields:
            db_v = _value(existing, f, core)
            new_v = _value(incoming, f, core)
            ws.cell(row_idx, col, "" if db_v is None else db_v).fill = DB_FILL
            ws.cell(row_idx, col + 1, "" if new_v is None else new_v).fill = NEW_FILL
            col += 2

    # ---- 列宽 ----
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 32
    for i in range(4, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = 18

    ws.freeze_panes = "D2"  # 冻结左 3 列 + 表头


def build_conflicts_xlsx(conflicts: list[dict[str, Any]]) -> bytes:
    """conflicts 列表 → xlsx 字节流。"""
    site_rows = [c for c in conflicts if c.get("kind") == "site"]
    lessor_rows = [c for c in conflicts if c.get("kind") == "lessor"]
    area_rows = [c for c in conflicts if c.get("kind") == "area"]
    road_rows = [c for c in conflicts if c.get("kind") == "road"]

    wb = Workbook()
    wb.remove(wb.active)  # 删默认空 sheet

    if site_rows:
        ws = wb.create_sheet("Site Conflicts")
        _write_sheet(ws, "site", site_rows)
    if road_rows:
        ws = wb.create_sheet("Road Conflicts")
        _write_sheet(ws, "road", road_rows)
    if lessor_rows:
        ws = wb.create_sheet("Lessor Conflicts")
        _write_sheet(ws, "lessor", lessor_rows)
    if area_rows:
        ws = wb.create_sheet("Area Conflicts")
        _write_sheet(ws, "area", area_rows)
    if not site_rows and not lessor_rows and not area_rows and not road_rows:
        ws = wb.create_sheet("Conflicts")
        ws.cell(1, 1, "无冲突").font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
