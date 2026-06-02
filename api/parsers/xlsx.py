"""Excel (Site 详表) 解析。

约定：
- 第 1 行：分类横幅 (BAISC / SAQ / TECHNICAL)，跳过
- 第 2 行：50 列真实字段名（PROJECT / SITE ID / OPTION / ... LATI / LONGI / ...）
- 第 3 行起：数据
"""

import io
from typing import Optional

from openpyxl import load_workbook

from .kml import ParseResult, SiteRow


class ParseError(Exception):
    pass


def _norm(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_float(v: object) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return None


def parse_xlsx(data: bytes) -> ParseResult:
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.worksheets[0]

    # #35：自动探测表头行（前 5 行内找含 SITE ID 的行）。
    # 旧模板第 1 行是分类横幅、表头落第 2 行；新模板无横幅、表头落第 1 行——都能认。
    # break 后生成器 rows 剩余即表头之后的数据行，下方数据循环原样不动。
    rows = ws.iter_rows(values_only=True)
    SCAN = 5
    header_row = None
    scanned = []
    for _ in range(SCAN):
        try:
            r = next(rows)
        except StopIteration:
            break
        scanned.append(r)
        norm = [(_norm(h) if h is not None else "") for h in r]
        if "SITE ID" in norm:
            header_row = r
            break
    if header_row is None:
        # 列名变体友好报错（如 Site_ID）：从扫描过的行里找候选
        candidates = []
        for r in scanned:
            for h in r:
                s = _norm(h)
                if s and "site" in s.lower() and "id" in s.lower():
                    candidates.append(s)
        raise ParseError(
            f"找不到表头行：前 {SCAN} 行内未发现 'SITE ID' 列"
            + (f"（检测到候选列：{', '.join(candidates)}）" if candidates else "")
        )

    headers = [(_norm(h) if h is not None else "") for h in header_row]

    sid_idx = headers.index("SITE ID")
    opt_idx = headers.index("OPTION") if "OPTION" in headers else None
    project_idx = headers.index("PROJECT") if "PROJECT" in headers else None
    status_idx = headers.index("SITE STATUS") if "SITE STATUS" in headers else None
    lati_idx = headers.index("LATI") if "LATI" in headers else None
    longi_idx = headers.index("LONGI") if "LONGI" in headers else None

    result = ParseResult()

    for row in rows:
        if row is None or all(c is None for c in row):
            continue
        site_id = _norm(row[sid_idx]) if sid_idx < len(row) else ""
        if not site_id:
            continue

        option = _norm(row[opt_idx]) if opt_idx is not None and opt_idx < len(row) else ""
        lati = _to_float(row[lati_idx]) if lati_idx is not None and lati_idx < len(row) else None
        longi = _to_float(row[longi_idx]) if longi_idx is not None and longi_idx < len(row) else None

        # F20 (V1.x #24/#25)：与 kml._SITE_CORE 对齐，排除盖戳三列源别名防重复显示
        _CORE = {
            "SITE ID", "OPTION", "PROJECT", "SITE STATUS", "LATI", "LONGI",
            "OPERATOR", "CATEGORY", "SITE CATEGORY", "TYPE", "SITE TYPE",
        }
        extras: dict[str, str] = {}
        for h, v in zip(headers, row):
            if not h or h in _CORE:
                continue
            sv = _norm(v)
            if sv:
                extras[h] = sv

        wkt = None
        if lati is not None and longi is not None:
            wkt = f"POINT({longi} {lati})"

        result.sites.append(
            SiteRow(
                site_id=site_id,
                option=option,
                project=_norm(row[project_idx]) or None if project_idx is not None else None,
                site_status=_norm(row[status_idx]) or None if status_idx is not None else None,
                lati=lati,
                longi=longi,
                extras=extras,
                wkt=wkt,
            )
        )

    return result
