"""F19 · 审计日志查询 / 导出端点（Spec V1.x #23）

- GET /api/audit-log         分页 + 筛选（action / from / to）
- GET /api/audit-log/export  按当前筛选导 Excel，导出本身再追写一条 audit_log_export

只读，不开 POST / DELETE / PATCH。
"""

import io
import json
from datetime import datetime
from typing import Any, Optional

from fastapi import APIRouter, Query, Request
from fastapi.responses import Response
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from audit import write_audit
from db import pool

router = APIRouter()

HEADER_FILL = PatternFill(start_color="FFEFEFEF", end_color="FFEFEFEF", fill_type="solid")
HEADER_FONT = Font(bold=True)
LEFT_TOP = Alignment(horizontal="left", vertical="top", wrap_text=True)


def _build_where(
    action: Optional[str], frm: Optional[str], to: Optional[str]
) -> tuple[str, list[Any]]:
    clauses: list[str] = []
    params: list[Any] = []
    if action:
        params.append(action)
        clauses.append(f"action = ${len(params)}")
    if frm:
        try:
            dt = datetime.fromisoformat(frm.replace("Z", "+00:00"))
            params.append(dt.replace(tzinfo=None))
            clauses.append(f"ts >= ${len(params)}")
        except ValueError:
            pass
    if to:
        try:
            dt = datetime.fromisoformat(to.replace("Z", "+00:00"))
            params.append(dt.replace(tzinfo=None))
            clauses.append(f"ts <= ${len(params)}")
        except ValueError:
            pass
    where = ("WHERE " + " AND ".join(clauses)) if clauses else ""
    return where, params


def _row_to_dict(r) -> dict[str, Any]:
    d: dict[str, Any] = {
        "id": r["id"],
        "ts": r["ts"].isoformat() if r["ts"] else None,
        "session_id": r["session_id"],
        "ip": r["ip"],
        "user_agent": r["user_agent"],
        "action": r["action"],
        "details": r["details"],
        "result": r["result"],
        "error_msg": r["error_msg"],
    }
    # asyncpg 把 jsonb 已经反序列化成 str，统一转 dict
    if isinstance(d["details"], str):
        try:
            d["details"] = json.loads(d["details"])
        except json.JSONDecodeError:
            pass
    return d


@router.get("/audit-log")
async def list_audit_log(
    action: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=500),
):
    where, params = _build_where(action, from_, to)
    offset = (page - 1) * page_size

    async with pool().acquire() as conn:
        total = await conn.fetchval(f"SELECT count(*) FROM audit_log {where}", *params)
        rows = await conn.fetch(
            f"SELECT * FROM audit_log {where} "
            f"ORDER BY ts DESC, id DESC "
            f"LIMIT {page_size} OFFSET {offset}",
            *params,
        )

    return {
        "items": [_row_to_dict(r) for r in rows],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get("/audit-log/export")
async def export_audit_log(
    request: Request,
    action: Optional[str] = Query(None),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str] = Query(None),
):
    where, params = _build_where(action, from_, to)
    async with pool().acquire() as conn:
        rows = await conn.fetch(
            f"SELECT * FROM audit_log {where} ORDER BY ts DESC, id DESC", *params
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Log"
    headers = ["ID", "时间 (ts)", "操作 (action)", "结果", "IP",
               "Session ID", "User Agent", "Details (JSON)", "Error"]
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = LEFT_TOP

    for i, r in enumerate(rows, start=2):
        details = r["details"]
        if isinstance(details, str):
            details_text = details
        else:
            details_text = json.dumps(details or {}, ensure_ascii=False, default=str)
        ws.cell(row=i, column=1, value=r["id"])
        ws.cell(row=i, column=2, value=r["ts"].isoformat() if r["ts"] else None)
        ws.cell(row=i, column=3, value=r["action"])
        ws.cell(row=i, column=4, value=r["result"])
        ws.cell(row=i, column=5, value=r["ip"])
        ws.cell(row=i, column=6, value=r["session_id"])
        ws.cell(row=i, column=7, value=r["user_agent"])
        ws.cell(row=i, column=8, value=details_text)
        ws.cell(row=i, column=9, value=r["error_msg"])

    widths = [8, 22, 30, 10, 18, 36, 50, 80, 30]
    for c, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    fname = f"audit_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # 元审计：导出本身再写一条（雷 7：必须有）
    await write_audit(
        action="audit_log_export",
        details={
            "file_name": fname,
            "exported_rows": len(rows),
            "filters": {"action": action, "from": from_, "to": to},
            "bytes": len(data),
        },
        request=request,
    )

    return Response(
        content=data,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{fname}"',
            "X-Filename": fname,
        },
    )
