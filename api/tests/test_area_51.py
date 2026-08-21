"""#51 Phase 17 · AREA 运营商区域面图层 回归测试（不连真库，monkeypatch 范式）。

覆盖：
- 解析：真实 NCR_BCA.kmz（158 面 / Name→name / extras）+ MultiGeometry 解壳
        + <Data name="Name"> → name + 无 SchemaData 不炸 + #area schema 回环
- 几何护栏：点/线导入 area 层被拒（跳过并报告）
- 冲突：同 operator+name 冲突 / 跨 operator 同名不冲突（SQL 带 operator 维度）
- scope：Globe PM 只见 Globe area / 类别级节点不授予 area / 无权限导入 403
- 导出自反：#area schema + 运营商分色 style（aabbggrr 字节序）+ 重导入全冲突
- 清洗：质心在海里标丢弃（classify_geoms mock countries + 决策应用）

范式同 test_site_crud_48.py / test_scope_filter_50.py：直调 handler + monkeypatch。
"""

import asyncio
import io
import zipfile
from pathlib import Path
from types import SimpleNamespace

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from areas import router as areas
from areas.router import list_areas
from auth.scopes import area_scope_operators, validate_scope_node
from exports import router as exports
from exporters.conflicts_xlsx import build_conflicts_xlsx
from exporters.kmz import build_kml
from imports import router as imports
from imports.cleaning import classify_geoms
from imports.router import ProceedBody, import_file, proceed_to_conflicts
from parsers.kml import parse_kml
from core import session_store

FIXTURE_KMZ = Path(__file__).parent / "fixtures" / "NCR_BCA.kmz"

KML_NS = 'xmlns="http://www.opengis.net/kml/2.2"'


# ---------- 公共假件（同 test_scope_filter_50.py 范式）----------


class _FakeRequest:
    def __init__(self, user=None):
        self.headers = {}
        self.client = None

        class _State:
            session_id = "test-sid"

        self.state = _State()
        if user is not None:
            self.state.user = user


def _user(scopes, is_admin=False, perms=None):
    return {"user_id": 9, "username": "u", "is_admin": is_admin,
            "perms": perms or {}, "scopes": scopes}


class _Acquire:
    def __init__(self, conn):
        self.conn = conn

    async def __aenter__(self):
        return self.conn

    async def __aexit__(self, *a):
        return False


class _Pool:
    def __init__(self, conn):
        self.conn = conn

    def acquire(self):
        return _Acquire(self.conn)


class _Tx:
    async def __aenter__(self):
        return self

    async def __aexit__(self, *a):
        return False


class _Conn:
    """SQL 感知的假连接：fetch 按 SQL 内容分类返回；记录全部调用供断言。"""

    def __init__(self, *, fetch_map=None, fetchrow_ret=None):
        # fetch_map: list[(匹配子串, 返回行列表)]，按序首个命中生效
        self.fetch_map = fetch_map or []
        self.fetchrow_ret = fetchrow_ret
        self.fetches = []  # [(sql, args)]
        self.executed = []  # [(sql, args)]

    async def fetch(self, sql, *args):
        self.fetches.append((sql, args))
        for pat, rows in self.fetch_map:
            if pat in sql:
                return rows
        return []

    async def fetchrow(self, sql, *args):
        return self.fetchrow_ret

    async def fetchval(self, sql, *args):
        return None

    async def execute(self, sql, *args):
        self.executed.append((sql, args))
        return "OK"

    def transaction(self):
        return _Tx()


class _FakeUpload:
    def __init__(self, filename: str, data: bytes):
        self.filename = filename
        self._data = data

    async def read(self):
        return self._data


def _area_row(name="QC_NORTH", operator="Globe", extras=None):
    return {
        "id": 7,
        "name": name,
        "operator": operator,
        "extras": extras if extras is not None else {"polygon_id": "002-0019"},
        "geojson": '{"type":"Polygon","coordinates":[[[121.0,14.6],[121.1,14.6],[121.1,14.7],[121.0,14.6]]]}',
        "geom_kml": ("<Polygon><outerBoundaryIs><LinearRing><coordinates>"
                     "121.0,14.6 121.1,14.6 121.1,14.7 121.0,14.6"
                     "</coordinates></LinearRing></outerBoundaryIs></Polygon>"),
    }


# =====================================================================
# 解析层
# =====================================================================


def test_parse_ncr_bca_real_sample():
    """真实样例：NCR_BCA.kmz 的 doc.kml → 158 面、Name 提取、extras 保留。"""
    with zipfile.ZipFile(FIXTURE_KMZ) as z:
        data = z.read("doc.kml")
    r = parse_kml(data)
    assert len(r.areas) == 158
    # 无 schema 的纯面文件：不串型到 site/road/lessor
    assert r.sites == [] and r.roads == [] and r.lessors == []
    names = [a.name for a in r.areas]
    assert all(names), "每个 area 都必须有 name（去重键）"
    assert len(set(names)) == 158
    # extras：业务列 polygon_id / geozone_pr 保留；Name/operator 不进 extras
    a0 = r.areas[0]
    assert "polygon_id" in a0.extras
    assert "geozone_pr" in a0.extras
    assert "Name" not in a0.extras and "name" not in a0.extras
    assert "operator" not in a0.extras and "OPERATOR" not in a0.extras
    # 面几何全部产出 WKT
    assert all(a.wkt and a.wkt.startswith("POLYGON((") for a in r.areas)


def test_parse_multigeometry_unwrapped():
    """MultiGeometry 壳内的 Polygon 被解壳提取。"""
    kml = f"""<?xml version="1.0" encoding="utf-8"?>
<kml {KML_NS}><Document>
  <Placemark>
    <ExtendedData>
      <Data name="Name"><value>MG_AREA</value></Data>
    </ExtendedData>
    <MultiGeometry>
      <Polygon><outerBoundaryIs><LinearRing><coordinates>
        121.0,14.6 121.1,14.6 121.1,14.7 121.0,14.6
      </coordinates></LinearRing></outerBoundaryIs></Polygon>
    </MultiGeometry>
  </Placemark>
</Document></kml>"""
    r = parse_kml(kml.encode())
    assert len(r.areas) == 1
    assert r.areas[0].name == "MG_AREA"
    assert r.areas[0].wkt.startswith("POLYGON((121.0 14.6")


def test_parse_multigeometry_picks_largest_polygon():
    """#51 review MEDIUM-1：MultiGeometry 含多个 Polygon 时取面积最大者（而非首个）。

    第一个是小三角（~0.005 平方度），第二个是大方块（1.0 平方度）→ 应取大方块。
    """
    kml = f"""<?xml version="1.0" encoding="utf-8"?>
<kml {KML_NS}><Document>
  <Placemark>
    <ExtendedData>
      <Data name="Name"><value>MG_BIG</value></Data>
    </ExtendedData>
    <MultiGeometry>
      <Polygon><outerBoundaryIs><LinearRing><coordinates>
        121.0,14.6 121.1,14.6 121.1,14.7 121.0,14.6
      </coordinates></LinearRing></outerBoundaryIs></Polygon>
      <Polygon><outerBoundaryIs><LinearRing><coordinates>
        120.0,10.0 121.0,10.0 121.0,11.0 120.0,11.0 120.0,10.0
      </coordinates></LinearRing></outerBoundaryIs></Polygon>
    </MultiGeometry>
  </Placemark>
</Document></kml>"""
    r = parse_kml(kml.encode())
    assert len(r.areas) == 1
    assert r.areas[0].wkt.startswith("POLYGON((120.0 10.0")  # 取到大的，不是首个小三角


def test_parse_data_name_and_extras_no_schemadata():
    """无 SchemaData 的外来文件：<Data name="Name"> → name，其余 Data 进 extras，空 value 跳过。"""
    kml = f"""<?xml version="1.0" encoding="utf-8"?>
<kml {KML_NS}><Document>
  <Placemark>
    <ExtendedData>
      <Data name="Name"><value>AREA_A</value></Data>
      <Data name="geozone_pr"><value>QUEZON CITY</value></Data>
      <Data name="empty_col"><value></value></Data>
    </ExtendedData>
    <Polygon><outerBoundaryIs><LinearRing><coordinates>
      121.0,14.6 121.1,14.6 121.1,14.7 121.0,14.6
    </coordinates></LinearRing></outerBoundaryIs></Polygon>
  </Placemark>
</Document></kml>"""
    r = parse_kml(kml.encode())
    assert len(r.areas) == 1
    a = r.areas[0]
    assert a.name == "AREA_A"
    assert a.extras == {"geozone_pr": "QUEZON CITY"}  # 空 value 不进 extras


def test_parse_no_extended_data_does_not_crash():
    """无 ExtendedData / 无 SchemaData 的面要素不报错；无 name → 跳过（去重键缺失）。"""
    kml = f"""<?xml version="1.0" encoding="utf-8"?>
<kml {KML_NS}><Document>
  <Placemark>
    <Polygon><outerBoundaryIs><LinearRing><coordinates>
      121.0,14.6 121.1,14.6 121.1,14.7 121.0,14.6
    </coordinates></LinearRing></outerBoundaryIs></Polygon>
  </Placemark>
</Document></kml>"""
    r = parse_kml(kml.encode())  # 不炸即过
    assert r.areas == []


def test_parse_area_schema_roundtrip_name():
    """本平台导出的 #area（SchemaData/SimpleData）→ name 精确回环，operator 不进 extras。"""
    kml = f"""<?xml version="1.0" encoding="utf-8"?>
<kml {KML_NS}><Document>
  <Schema name="area" id="area">
    <SimpleField name="name" type="string"></SimpleField>
    <SimpleField name="operator" type="string"></SimpleField>
    <SimpleField name="polygon_id" type="string"></SimpleField>
  </Schema>
  <Placemark>
    <ExtendedData>
      <SchemaData schemaUrl="#area">
        <SimpleData name="name">QC_NORTH</SimpleData>
        <SimpleData name="operator">Globe</SimpleData>
        <SimpleData name="polygon_id">002-0019</SimpleData>
      </SchemaData>
    </ExtendedData>
    <Polygon><outerBoundaryIs><LinearRing><coordinates>
      121.0,14.6 121.1,14.6 121.1,14.7 121.0,14.6
    </coordinates></LinearRing></outerBoundaryIs></Polygon>
  </Placemark>
</Document></kml>"""
    r = parse_kml(kml.encode())
    assert len(r.areas) == 1
    a = r.areas[0]
    assert a.name == "QC_NORTH"
    assert a.extras == {"polygon_id": "002-0019"}  # name/operator 是保留列


# =====================================================================
# 几何护栏：点/线导入 area 层被拒
# =====================================================================


def _guard_kml() -> bytes:
    return f"""<?xml version="1.0" encoding="utf-8"?>
<kml {KML_NS}><Document>
  <Placemark>
    <ExtendedData><SchemaData schemaUrl="#site">
      <SimpleData name="SITE ID">S1</SimpleData>
      <SimpleData name="OPTION">A</SimpleData>
    </SchemaData></ExtendedData>
    <Point><coordinates>121.0,14.6</coordinates></Point>
  </Placemark>
  <Placemark>
    <ExtendedData><SchemaData schemaUrl="#road">
      <SimpleData name="Property">R1</SimpleData>
    </SchemaData></ExtendedData>
    <LineString><coordinates>121.0,14.6 121.1,14.7</coordinates></LineString>
  </Placemark>
  <Placemark>
    <ExtendedData>
      <Data name="Name"><value>AREA_OK</value></Data>
    </ExtendedData>
    <Polygon><outerBoundaryIs><LinearRing><coordinates>
      121.0,14.6 121.1,14.6 121.1,14.7 121.0,14.6
    </coordinates></LinearRing></outerBoundaryIs></Polygon>
  </Placemark>
</Document></kml>""".encode()


def test_import_area_guard_rejects_point_and_line(monkeypatch):
    """target_kind=area：点/线要素被护栏跳过并报告，面正常入池。"""
    conn = _Conn(fetchrow_ret=None)  # baseline_state 空 → baseline None
    monkeypatch.setattr(imports, "pool", lambda: _Pool(conn))
    req = _FakeRequest(user=_user(["site:Globe"], perms={"import": True}))
    resp = asyncio.run(import_file(
        _FakeUpload("areas.kml", _guard_kml()), req,
        operator="Globe", category=None, type_=None, target_kind="area",
    ))
    guard = resp["geometry_guard"]
    assert guard["skipped"]["site"] == 1  # 点被拒
    assert guard["skipped"]["road"] == 1  # 线被拒
    assert guard["skipped"]["lessor"] == 0
    assert resp["file"]["parsed"]["area"] == 1
    assert resp["summary"]["after_dedup"]["area"] == 1
    # 点/线没有混进解析计数
    assert resp["file"]["parsed"]["site"] == 0
    assert resp["file"]["parsed"]["road"] == 0


def test_import_area_requires_operator_stamp():
    """area 导入盖戳 operator 必填（缺 → 403，同 F20 盖戳模型）。"""
    req = _FakeRequest(user=_user(["site:Globe"], perms={"import": True}))
    with pytest.raises(HTTPException) as exc:
        asyncio.run(import_file(
            _FakeUpload("areas.kml", _guard_kml()), req,
            operator=None, category=None, type_=None, target_kind="area",
        ))
    assert exc.value.status_code == 403


# =====================================================================
# 冲突：operator+name 三路径
# =====================================================================


def _make_import_session(area_pool, stamp_operator="Globe"):
    return session_store.create(
        {
            "file_name": "areas.kml",
            "site_pool": {},
            "lessor_pool": {},
            "road_pool": [],
            "area_pool": area_pool,
            "cleanings": [],
            "baseline_region": None,
            "target_kind": "area",
            "stamp_operator": stamp_operator,
            "stamp_category": None,
            "stamp_type": None,
        },
        state="cleaning",
    )


def _pool_row(name, operator="Globe", wkt="POLYGON((121.0 14.6, 121.1 14.6, 121.1 14.7, 121.0 14.6))"):
    return {"name": name, "extras": {}, "wkt": wkt, "source_file": "areas.kml"}


def test_area_conflict_same_operator_same_name(monkeypatch):
    """同 operator + 同 name → 冲突（kind=area，key 带盖戳运营商）。"""
    conn = _Conn(fetch_map=[
        ("FROM area", [{"id": 1, "name": "QC_NORTH", "operator": "Globe", "extras": {}}]),
    ])
    monkeypatch.setattr(imports, "pool", lambda: _Pool(conn))
    sid = _make_import_session({"qc_north": _pool_row("QC_NORTH")})
    resp = asyncio.run(proceed_to_conflicts(sid, ProceedBody()))
    assert resp["summary"]["area"] == {"non_conflict": 0, "conflict": 1}
    c = resp["conflicts"][0]
    assert c["kind"] == "area"
    assert c["key"] == "area:Globe:QC_NORTH"
    assert c["existing"]["name"] == "QC_NORTH"
    # area 冲突查询带 operator 维度
    area_fetch = [f for f in conn.fetches if "FROM area" in f[0]]
    assert len(area_fetch) == 1
    assert "WHERE operator = $1" in area_fetch[0][0]
    assert area_fetch[0][1] == ("Globe",)


def test_area_cross_operator_same_name_no_conflict(monkeypatch):
    """跨 operator 同名不冲突：SQL 只查盖戳运营商，Smart 的同名面不进结果 → 新行。"""
    # DB 里 (Smart, QC_NORTH) 存在，但盖戳 Globe → area 查询返回空（operator 过滤生效）
    conn = _Conn(fetch_map=[("FROM area", [])])
    monkeypatch.setattr(imports, "pool", lambda: _Pool(conn))
    sid = _make_import_session({"qc_north": _pool_row("QC_NORTH")}, stamp_operator="Globe")
    resp = asyncio.run(proceed_to_conflicts(sid, ProceedBody()))
    assert resp["summary"]["area"] == {"non_conflict": 1, "conflict": 0}
    area_fetch = [f for f in conn.fetches if "FROM area" in f[0]]
    assert area_fetch[0][1] == ("Globe",)  # 只按盖戳运营商查


# =====================================================================
# 冲突 Excel：area 冲突行不得被静默丢弃（#51 review CRITICAL-1）
# =====================================================================


def _area_conflict(name="QC_NORTH"):
    """与 proceed_to_conflicts 产出同构的 area 冲突 dict。"""
    return {
        "key": f"area:Globe:{name}",
        "kind": "area",
        "name": name,
        "existing": {"id": 1, "name": name, "operator": "Globe",
                     "extras": {"polygon_id": "002-0019"}},
        "incoming": {"name": name, "extras": {"polygon_id": "002-0020"},
                     "wkt": "POLYGON((121.0 14.6, 121.1 14.6, 121.1 14.7, 121.0 14.6))",
                     "source_file": "areas.kml"},
        "source_file": "areas.kml",
    }


def test_conflicts_xlsx_pure_area_not_lied_as_no_conflict():
    """纯 area 冲突 → 有 Area Conflicts sheet，A1 不是「无冲突」谎报。"""
    data = build_conflicts_xlsx([_area_conflict()])
    wb = load_workbook(io.BytesIO(data))
    assert "Area Conflicts" in wb.sheetnames
    ws = wb["Area Conflicts"]
    assert ws.cell(1, 1).value != "无冲突"
    # 核心列 name/operator + extras 展开；name/operator 不回灌进 extras 列
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "name [DB]" in headers and "operator [DB]" in headers
    assert "polygon_id [DB]" in headers and "polygon_id [新]" in headers
    # 数据行：类型/名称/来源文件 + DB 值
    assert ws.cell(2, 1).value == "area"
    assert ws.cell(2, 2).value == "QC_NORTH"
    assert ws.cell(2, 3).value == "areas.kml"
    db_name_col = headers.index("name [DB]") + 1
    assert ws.cell(2, db_name_col).value == "QC_NORTH"
    db_pid_col = headers.index("polygon_id [DB]") + 1
    assert ws.cell(2, db_pid_col).value == "002-0019"
    assert ws.cell(2, db_pid_col + 1).value == "002-0020"  # [新]


def test_conflicts_xlsx_mixed_site_and_area():
    """site + area 混合冲突 → 两个 sheet 各归各位（area 不再被丢）。"""
    site_conflict = {
        "key": "site:S1:A", "kind": "site", "name": "S1 / A",
        "existing": {"site_id": "S1", "option": "A", "project": None,
                     "site_status": None, "lati": 14.6, "longi": 121.0, "extras": {}},
        "incoming": {"site_id": "S1", "option": "A", "project": None,
                     "site_status": None, "lati": 14.7, "longi": 121.0,
                     "extras": {}, "source_file": "sites.kml"},
        "source_file": "sites.kml",
    }
    data = build_conflicts_xlsx([site_conflict, _area_conflict()])
    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Site Conflicts", "Area Conflicts"]
    assert wb["Site Conflicts"].cell(2, 2).value == "S1 / A"
    assert wb["Area Conflicts"].cell(2, 2).value == "QC_NORTH"


# =====================================================================
# 冲突 Excel：road 冲突行不得被静默丢弃（#51 area 同根缺陷的 F20 road 变体）
# =====================================================================


def _road_conflict(prop="EDSA-SEG1"):
    """与 proceed_to_conflicts 产出同构的 road 冲突 dict（F20：按 Property 判重）。"""
    return {
        "key": f"road:{prop}",
        "kind": "road",
        "name": prop,
        "existing": {"id": 7, "property": prop,
                     "extras": {"Length": "1.2km"}, "source_file": "roads.kml"},
        "incoming": {"property": prop, "extras": {"Length": "1.5km"},
                     "wkt": "LINESTRING(121.0 14.6, 121.1 14.7)",
                     "source_file": "roads.kml"},
        "source_file": "roads.kml",
    }


def test_conflicts_xlsx_pure_road_not_lied_as_no_conflict():
    """纯 road 冲突 → 有 Road Conflicts sheet，A1 不是「无冲突」谎报。"""
    data = build_conflicts_xlsx([_road_conflict()])
    wb = load_workbook(io.BytesIO(data))
    assert "Road Conflicts" in wb.sheetnames
    ws = wb["Road Conflicts"]
    assert ws.cell(1, 1).value != "无冲突"
    # 核心列 property（去重键）+ extras 展开；Property 不回灌进 extras 列
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    assert "property [DB]" in headers and "property [新]" in headers
    assert "Length [DB]" in headers and "Length [新]" in headers
    # 数据行：类型/名称/来源文件 + DB/新值
    assert ws.cell(2, 1).value == "road"
    assert ws.cell(2, 2).value == "EDSA-SEG1"
    assert ws.cell(2, 3).value == "roads.kml"
    db_prop_col = headers.index("property [DB]") + 1
    assert ws.cell(2, db_prop_col).value == "EDSA-SEG1"
    db_len_col = headers.index("Length [DB]") + 1
    assert ws.cell(2, db_len_col).value == "1.2km"
    assert ws.cell(2, db_len_col + 1).value == "1.5km"  # [新]


# =====================================================================
# scope：areas 列表过滤 + 导入 403
# =====================================================================


def test_area_scope_operators_grant_rules():
    """授予口径：site:<op> / site:<op>:AREA 授予；类别级节点不授予；根/* 全量。"""
    assert area_scope_operators(["*"]) is None
    assert area_scope_operators(["site"]) is None
    assert area_scope_operators(["site:Globe"]) == {"Globe"}
    assert area_scope_operators(["site:Globe:AREA"]) == {"Globe"}
    assert area_scope_operators(["site:Globe:SURVEY"]) == set()  # 类别级不授予 area
    assert area_scope_operators(["site:Globe:AREA", "site:Smart"]) == {"Globe", "Smart"}


def test_validate_scope_node_area():
    assert validate_scope_node("site:Globe:AREA") is True
    assert validate_scope_node("site:Dito:AREA") is True
    assert validate_scope_node("site:Globe:BAD") is False
    assert validate_scope_node("AREA") is False


def test_list_areas_globe_pm_only_globe(monkeypatch):
    """Globe PM（site:Globe:AREA）→ SQL 带 operator = ANY，仅 Globe 行。"""
    conn = _Conn(fetch_map=[("FROM area", [_area_row("QC_NORTH", "Globe")])])
    monkeypatch.setattr(areas, "pool", lambda: _Pool(conn))
    req = _FakeRequest(user=_user(["site:Globe:AREA"], perms={}))
    resp = asyncio.run(list_areas(req))
    sql, args = conn.fetches[0]
    assert "operator = ANY($1::text[])" in sql
    assert args == (["Globe"],)
    assert resp["type"] == "FeatureCollection"
    assert len(resp["features"]) == 1
    f = resp["features"][0]
    assert f["properties"]["kind"] == "area"
    assert f["properties"]["name"] == "QC_NORTH"
    assert f["properties"]["operator"] == "Globe"
    assert f["properties"]["polygon_id"] == "002-0019"  # extras 展开
    assert f["geometry"]["type"] == "Polygon"


def test_list_areas_category_node_sees_nothing_without_db(monkeypatch):
    """只有类别级节点（site:Globe:SURVEY）→ area 空集，不查库。"""
    monkeypatch.setattr(
        areas, "pool",
        lambda: (_ for _ in ()).throw(AssertionError("不应查库")),
    )
    req = _FakeRequest(user=_user(["site:Globe:SURVEY"], perms={}))
    resp = asyncio.run(list_areas(req))
    assert resp == {"type": "FeatureCollection", "features": []}


def test_list_areas_admin_full(monkeypatch):
    """admin → 全量，无 operator WHERE。"""
    conn = _Conn(fetch_map=[("FROM area", [_area_row("A", "Globe"), _area_row("B", "Smart")])])
    monkeypatch.setattr(areas, "pool", lambda: _Pool(conn))
    req = _FakeRequest(user=_user([], is_admin=True))
    resp = asyncio.run(list_areas(req))
    sql, _ = conn.fetches[0]
    assert "operator = ANY" not in sql
    assert len(resp["features"]) == 2


def test_import_area_out_of_scope_403():
    """Globe 盖戳但只见 Smart → 403。"""
    req = _FakeRequest(user=_user(["site:Smart"], perms={"import": True}))
    with pytest.raises(HTTPException) as exc:
        asyncio.run(import_file(
            None, req, operator="Globe", category=None, type_=None, target_kind="area",
        ))
    assert exc.value.status_code == 403


def test_import_area_area_node_scope_passes():
    """site:Globe:AREA 授予导入权限 → 过 scope 校验（走到文件类型检测报 400 而非 403）。"""
    req = _FakeRequest(user=_user(["site:Globe:AREA"], perms={"import": True}))
    with pytest.raises(HTTPException) as exc:
        asyncio.run(import_file(
            SimpleNamespace(filename=""), req,
            operator="Globe", category=None, type_=None, target_kind="area",
        ))
    assert exc.value.status_code == 400  # 文件名空 → 400，证明 scope 已过


# =====================================================================
# 导出自反：#area schema + 分色 style + 重导入全冲突
# =====================================================================


def test_export_area_schema_and_operator_styles():
    """#area schema 存在；按运营商分色（aabbggrr 字节序，alpha≈35%=0x59）。"""
    rows = [
        _area_row("A1", "Globe"),
        _area_row("A2", "Smart"),
        _area_row("A3", "Dito"),
        _area_row("A4", "OtherOp"),
    ]
    kml = build_kml([], [], [], area_rows=rows)
    assert '<Schema name="area" id="area">' in kml
    assert 'schemaUrl="#area"' in kml
    assert "Area Library" in kml
    # 分色 style：填充 59（~35% 透明）+ aabbggrr
    assert "59F6823B" in kml  # Globe #3b82f6
    assert "595EC522" in kml  # Smart #22c55e
    assert "594444EF" in kml  # Dito  #ef4444
    assert "#poly-area-globe" in kml
    assert "#poly-area-smart" in kml
    assert "#poly-area-dito" in kml
    assert "#poly-area-other" in kml  # 未知运营商兜底
    # name/operator 作 SimpleData 写出
    assert '<SimpleData name="name">A1</SimpleData>' in kml
    assert '<SimpleData name="operator">Globe</SimpleData>' in kml
    assert '<SimpleData name="polygon_id">002-0019</SimpleData>' in kml


def test_export_empty_area_rows_backward_compatible():
    """area_rows 缺省/空 → 不产出 #area schema 与 Area Library（向后兼容三类导出）。"""
    kml = build_kml([], [], [])
    assert '<Schema name="area"' not in kml
    assert "Area Library" not in kml


def test_export_reimport_full_conflict(monkeypatch):
    """自反契约：导出 → parse_kml 回读 name 精确回环 → 重导入 100% 命中冲突。"""
    rows = [_area_row("QC_NORTH", "Globe"), _area_row("QC_SOUTH", "Globe")]
    kml = build_kml([], [], [], area_rows=rows)
    parsed = parse_kml(kml.encode())
    assert [a.name for a in parsed.areas] == ["QC_NORTH", "QC_SOUTH"]
    assert parsed.sites == [] and parsed.roads == [] and parsed.lessors == []

    # 重导入：同 operator 下同名 → 全部冲突
    conn = _Conn(fetch_map=[
        ("FROM area", [
            {"id": 1, "name": "QC_NORTH", "operator": "Globe", "extras": {}},
            {"id": 2, "name": "QC_SOUTH", "operator": "Globe", "extras": {}},
        ]),
    ])
    monkeypatch.setattr(imports, "pool", lambda: _Pool(conn))
    pool_rows = {
        imports._area_key(a.name): {
            "name": a.name, "extras": a.extras, "wkt": a.wkt, "source_file": "re.kmz",
        }
        for a in parsed.areas
    }
    sid = _make_import_session(pool_rows)
    resp = asyncio.run(proceed_to_conflicts(sid, ProceedBody()))
    assert resp["summary"]["area"] == {"non_conflict": 0, "conflict": 2}


def test_export_counts_include_area():
    """审计 counts 含 area。"""
    _, _, counts = exports._build_kmz_meta("full", {
        "site": [], "road": [], "lessor": [], "area": [_area_row()],
    })
    assert counts == {"site": 0, "road": 0, "lessor": 0, "area": 1}


# =====================================================================
# 清洗：质心判定（mock countries）
# =====================================================================


def test_classify_geoms_centroid_in_sea():
    """质心落海（countries 无命中 → iso None）→ in_sea=True。"""
    conn = _Conn(fetch_map=[
        ("centroids", [{
            "row_id": "area:sea_poly",
            "country_iso_a2": None,
            "country_name_zh": None,
            "country_name_en": None,
        }]),
    ])
    out = asyncio.run(classify_geoms(
        conn,
        [{"row_id": "area:sea_poly",
          "wkt": "POLYGON((120.0 10.0, 120.1 10.0, 120.1 10.1, 120.0 10.0))"}],
        "PH",
    ))
    assert out["area:sea_poly"]["in_sea"] is True
    assert out["area:sea_poly"]["not_in_baseline"] is False
    # SQL 走的是 ST_Centroid 质心口径（与选区导出一致）
    sql, _ = conn.fetches[0]
    assert "ST_Centroid" in sql


def test_classify_geoms_centroid_not_in_baseline():
    """质心落非基准国 → not_in_baseline=True。"""
    conn = _Conn(fetch_map=[
        ("centroids", [{
            "row_id": "area:foreign",
            "country_iso_a2": "MY",
            "country_name_zh": "马来西亚",
            "country_name_en": "Malaysia",
        }]),
    ])
    out = asyncio.run(classify_geoms(
        conn, [{"row_id": "area:foreign", "wkt": "POLYGON((1 1, 2 1, 2 2, 1 1))"}], "PH",
    ))
    assert out["area:foreign"]["in_sea"] is False
    assert out["area:foreign"]["not_in_baseline"] is True


def test_proceed_cleaning_area_in_sea_discard(monkeypatch):
    """area 清洗决策：in_sea 默认 discard → 从 pool 剔除；显式 keep → 保留。"""
    conn = _Conn(fetch_map=[("FROM area", [])])
    monkeypatch.setattr(imports, "pool", lambda: _Pool(conn))
    sid = _make_import_session({
        "sea_poly": _pool_row("SEA_POLY"),
        "keep_poly": _pool_row("KEEP_POLY"),
    })
    session_store.update(sid, {
        "cleanings": [
            {"row_id": "area:sea_poly", "kind": "area", "name": "SEA_POLY",
             "file_name": "areas.kml", "issue": "in_sea", "current_coord": None,
             "fixed_coord_preview": None, "default_action": "discard",
             "country_iso_a2": None},
            {"row_id": "area:keep_poly", "kind": "area", "name": "KEEP_POLY",
             "file_name": "areas.kml", "issue": "in_sea", "current_coord": None,
             "fixed_coord_preview": None, "default_action": "discard",
             "country_iso_a2": None},
        ],
    })
    body = ProceedBody(decisions=[
        imports.CleaningDecision(row_id="area:keep_poly", action="keep"),
    ])
    resp = asyncio.run(proceed_to_conflicts(sid, body))
    # sea_poly 默认 discard 剔除；keep_poly 显式 keep 保留 → 非冲突新行
    assert resp["summary"]["area"] == {"non_conflict": 1, "conflict": 0}
    assert resp["cleaning_stats"]["discarded"] == 1
    assert resp["cleaning_stats"]["kept"] == 1
    s = session_store.get(sid)
    assert list(s["area_pool_cleaned"].keys()) == ["keep_poly"]
